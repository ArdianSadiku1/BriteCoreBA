
#import libraries
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle

#################################################
# I. Load and parse data using pandas           #
#################################################

#load files into pandas DataFrame
p_DF= pd.read_csv("data/policies.csv")
f_DF = pd.read_csv('data/fees.csv')
ph_DF= pd.read_csv('data/policyholders.csv')
r_DF = pd.read_csv('data/revisions.csv')


#Concatenate the policy holders by revisionID
ph_DF=ph_DF.groupby('revisionId').agg(lambda x: x.tolist()).reset_index()
ph_DF['policyholderName'] = ph_DF.policyholderName.apply(', '.join)


#Remove un-reportable data "Open and Pending The data associated with the revision is in progress and, therefore, is neither in effect nor reportable")
# This will filter out the unecessary data and improve the performance & resource usage of the script (not noticable on this case , but yes when dealing with huge amount of data)
r_DF=r_DF.loc[(r_DF['revisionState']!='open') & (r_DF['revisionState']!='pending')]

#Convert dates from string to date
r_DF['revisionDate'] = pd.to_datetime(r_DF['revisionDate'])
r_DF['commitDate'] = pd.to_datetime(r_DF['commitDate'])
r_DF['cancelDate'] = pd.to_datetime(r_DF['cancelDate'])
r_DF['effectiveDate'] = pd.to_datetime(r_DF['effectiveDate'])

#Join Policies & Policy Revisions, on policyID, we do a left join from revisions to policies since one policy can have multiple reviosions; we could do the opposite with a right join
pr_DF = pd.merge(r_DF, p_DF, how = 'left', left_on='policyId', right_on='policyId' )

#join Policies & POlicy Revisions with policy Holders
prh_DF = pd.merge(pr_DF, ph_DF, how = 'inner', left_on='revisionId', right_on='revisionId')

# Remove rows from fees without values
#f_DF.dropna(subset=['writtenFee'], inplace=True)

#Sum up fees for by revision
f_DF=f_DF.groupby('revisionId')['writtenFee'].sum().reset_index()

# Join the newly created DataFrame with the Fees DataFrame
prhf_DF = pd.merge(prh_DF, f_DF, how = 'inner', left_on='revisionId', right_on='revisionId')

#The Policy Fees need to be removed from flat cancellations. These are policy cancellation 
#transactions where the cancel date is less or equal than the Effective Date.
prhf_DF['writtenFee'].fillna(0, inplace=True)
mask = (prhf_DF.cancelDate <= prhf_DF.effectiveDate) & (prhf_DF.policyStatus=='Canceled')
prhf_DF.loc[mask, 'writtenFee'] = 0


#Calculate the change in Premium for each policy
prhf_DF['Change_in_Premium'] = 0
prhf_DF.set_index(['policyNumber','createDate'], inplace=True)
prhf_DF.sort_index(inplace=True)
prhf_DF['Change_in_Premium'] = \
prhf_DF.groupby(['policyNumber'])['writtenPremium'].transform(lambda x: x.diff()) 

#reset the index the NaN's on the DataFRame
prhf_DF=prhf_DF.reset_index()
prhf_DF['Change_in_Premium'].fillna(0, inplace=True)

#Filter out only the columns needed for the excel file
ExcelCols_DF = prhf_DF.loc[:, \
['policyNumber','policyholderName', 'policyStatus', 'effectiveDate','Change_in_Premium', 'writtenFee']]

#########################################################################################
#  II. Create the excel file                                                            #
#########################################################################################

wb = Workbook()
ws = wb.active

ws.cell(row=1, column=1).value = 'Insurio, Inc'
ws.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=False)
ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)

ws.cell(row=2, column=1).value = 'June 2018 Premium Report'
ws.cell(row=2, column=1).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=False)
ws.cell(row=2, column=1).font = openpyxl.styles.Font(bold=True)

#Shift 4 row down for the table
ws.cell(row=4, column=1)

for r in dataframe_to_rows(ExcelCols_DF, index=False, header=True):
    ws.append(r)
    
#Columns
ColumH=(list(ExcelCols_DF.columns))
#rename the columns
ws.cell(row=5, column=1).value = 'Policy Number'
ws.cell(row=5, column=2).value = 'Named Insured'
ws.cell(row=5, column=3).value = 'Transaction Type'
ws.cell(row=5, column=4).value = 'Effective Date'
ws.cell(row=5, column=5).value = 'Change in Premiums'
ws.cell(row=5, column=6).value = 'Policy Fees'

#add style bold and underline to header row
for c in range (1, len(ColumH)+1):
    ws.cell(row=5, column=c).font = openpyxl.styles.Font(bold=True, underline='single')
    ws.cell(row=5, column=c).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor='A9D08E')

#Format Columns

#Date
dateFormat = NamedStyle(name='date', number_format="M/D/YYYY")
for col in {4}:
    for row in range (6,ws.max_row+1):
        ws.cell(row=row,column=col).style = dateFormat

#Currency
currencyFormat = NamedStyle(name='currency', number_format=r'_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)')
for col in {5, 6}:  
    for row in range (6,ws.max_row+1):
        ws.cell(row=row,column=col).style = currencyFormat

for col in ws.columns:
    max_length = 8.43 # the default column width
    column = col[0].column 
    for cell in col:
        if cell.coordinate in ws.merged_cells: # Skip merged cells
            continue
        try: # If the cell has content
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 4) 
    ws.column_dimensions[column].width = adjusted_width

#add filter    
ws.auto_filter.ref = 'A5:F5'

#Freeze cells at row 6
ws.freeze_panes = ws['A6']

wb.save("Insurio_Inc_Premium_Report_2018_06.xlsx")